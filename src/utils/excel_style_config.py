from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ===== 基础样式 =====
HEADER_FILL = PatternFill(patternType="solid", fgColor="FF3D5DD0")
DATA_FILL = PatternFill(patternType="solid", fgColor="FFE8EAF7")

TITLE_FONT_18 = Font(name="思源黑体 CN Regular", size=18, bold=False, color="FFFFFFFF")
HEADER_FONT_12 = Font(name="思源黑体 CN Regular", size=12, bold=True, color="FFFFFFFF")
HEADER_FONT_105 = Font(name="思源黑体 CN Regular", size=10.5, bold=True, color="FFFFFFFF")
HEADER_FONT_10 = Font(name="思源黑体 CN Regular", size=10, bold=True, color="FFFFFFFF")
DATA_FONT_12 = Font(name="思源黑体 CN Regular", size=12, bold=False, color="FF000000")

ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrapText=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_VCENTER = Alignment(vertical="center")

PERCENT_FORMAT = "0.00%"
DATE_FORMAT = "mm-dd-yy"
SIZE_FORMAT = "0.00_"

# 边框（细白色）
WHITE_THIN_BORDER = Border(
    left=Side(style="thin", color="FFFFFFFF"),
    right=Side(style="thin", color="FFFFFFFF"),
    top=Side(style="thin", color="FFFFFFFF"),
    bottom=Side(style="thin", color="FFFFFFFF"),
)


def apply_product_sheet_style(ws, max_row: int, max_col: int) -> None:
    """应用“产品及基准收益率”样式。"""
    # 合并 A1:A2
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    # A1 大标题样式（与样本一致）
    cell = ws.cell(row=1, column=1)
    cell.font = TITLE_FONT_18
    cell.fill = HEADER_FILL
    cell.alignment = ALIGN_CENTER_WRAP

    # 第一行：期间标题
    for col in range(2, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT_12
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER_WRAP

    # 第二行：日期区间
    for col in range(2, max_col + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = HEADER_FONT_105
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER_WRAP

    # 左侧行标（第3行开始）
    for row in range(3, max_row + 1):
        cell = ws.cell(row=row, column=1)
        cell.font = HEADER_FONT_12
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER_WRAP

    # 数据区域
    for row in range(3, max_row + 1):
        for col in range(2, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT_12
            cell.fill = DATA_FILL
            cell.alignment = ALIGN_CENTER
            cell.number_format = PERCENT_FORMAT
            cell.border = WHITE_THIN_BORDER

    # 为表格区域统一加边框（含表头和行标）
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = WHITE_THIN_BORDER


def apply_weekly_sheet_style(ws, max_row: int, max_col: int) -> None:
    """应用“周收益率曲线”样式。"""
    # 日期列格式
    for row in range(2, max_row + 1):
        cell = ws.cell(row=row, column=1)
        cell.alignment = ALIGN_VCENTER
        cell.number_format = DATE_FORMAT

    # 数值列格式
    for col in range(2, max_col + 1):
        header = ws.cell(row=1, column=col).value
        is_size = isinstance(header, str) and "规模" in header
        for row in range(2, max_row + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = ALIGN_VCENTER
            cell.number_format = SIZE_FORMAT if is_size else PERCENT_FORMAT


def apply_monthly_sheet_style(ws, max_row: int, max_col: int) -> None:
    """应用“月度收益率”样式。"""
    # 表头
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT_10
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER_WRAP

    # 左侧行标
    for row in range(2, max_row + 1):
        cell = ws.cell(row=row, column=1)
        cell.font = HEADER_FONT_10
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER_WRAP

    # 数据区域
    for row in range(2, max_row + 1):
        for col in range(2, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT_12
            cell.fill = DATA_FILL
            cell.alignment = ALIGN_CENTER
            cell.number_format = PERCENT_FORMAT
            cell.border = WHITE_THIN_BORDER

    # 为表格区域统一加边框（含表头和行标）
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = WHITE_THIN_BORDER
